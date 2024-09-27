import fitz
from fitz import Rect, Shape
import openpyxl
import re

green = (0.661997377872467, 0.8189975023269653, 0.5559929609298706)
blue = (0.0, 0.6899977326393127, 0.9409933686256409)

with fitz.open("input.pdf") as pdf:
    counter = 1
    text_list = [["", "", ""]]
    rects = []

    for page_num, page in enumerate(pdf):
        drawings = page.get_drawings()
        text = page.get_text()

        for drawing in drawings:
            if drawing["fill"] == green:
                rect = drawing["rect"]
                # rect[1] -= 15
                rects.append((page_num, rect, "heading"))
                page.add_highlight_annot(rect)
            # if drawing["fill"] == blue:
            #     rect = drawing["rect"]
            #     # rect[1] -= 15
            #     rects.append((page_num, rect, "subheading"))
            #     page.add_highlight_annot(rect)

    rects.sort(key=lambda x: (x[0], x[1][1]))

    prev_page_num = -1
    prev_rect_end = (0, 0)
    for page_num, rect, category in rects:
        page = pdf[prev_page_num]
        if prev_page_num != page_num:

            current_rect = (*prev_rect_end, *(pdf[prev_page_num].rect.bottom_right))
            text_list[-1][2] += pdf[prev_page_num].get_textbox(current_rect)
            page.draw_rect(current_rect, color=[0, 1, 1], overlay=False, width=3)

            prev_page_num += 1
            while prev_page_num != page_num:
                page = pdf[prev_page_num]
                page.draw_rect(page.rect, color=[0, 1, 1], width=10, overlay=False)
                text_list[-1][2] += pdf[prev_page_num].get_text()
                prev_page_num += 1
            page = pdf[prev_page_num]
            current_rect = Rect(0, 0, rect.top_right)
            text_list[-1][2] += page.get_textbox(current_rect)
            page.draw_rect(current_rect, color=[0, 1, 1], overlay=False, width=3)

        else:
            page = pdf[prev_page_num]
            current_rect = Rect(*prev_rect_end, *rect.top_right)
            text_list[-1][2] += page.get_textbox(current_rect)
            page.draw_rect(current_rect, color=[0, 1, 1], overlay=False, width=3)
            # Shape.draw_rect(current_rect,width=3)

        prev_content = text_list[-1][-1][::-1]
        match = re.search("\d+－\d+", prev_content)
        if match is not None:
            topic = prev_content[: match.end()][::-1]
            text_list[-1][2] = text_list[-1][2][: len(text_list[-1][2]) - match.end()]
        else:
            topic = ""
        text_list.append(
            [counter, topic + "\n" + page.get_textbox(rect).strip(), "", category]
        )
        prev_rect_end = rect.bottom_left
        counter += 1

    current_rect = (*prev_rect_end, *(pdf[prev_page_num].rect.bottom_right))
    text_list[-1][2] += page.get_textbox(current_rect)
    prev_page_num += 1
    while prev_page_num < pdf.page_count:
        text_list[-1][2] += page.get_text()
        prev_page_num += 1

    pdf.save("split.pdf")

workbook = openpyxl.Workbook()
sheet = workbook.active

for i, text in enumerate(text_list):
    if isinstance(text, tuple) or isinstance(text, list):
        if text[-1] in ["subheading", "heading"]:
            if text[-1] == "heading":
                for j, data in enumerate(text):
                    sheet.cell(row=i + 1, column=j + 1, value=data)
            else:
                for j, data in enumerate(text):
                    sheet.cell(row=i + 1, column=j + 2, value=data)
        else:
            for j, data in enumerate(text):
                sheet.cell(row=i + 1, column=j + 1, value=data)
    else:
        sheet.cell(row=i + 1, column=1, value=text)


workbook.save("sections.xlsx")
